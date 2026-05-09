#!/usr/bin/env python3
"""5000 sınav sonucundan Excel raporu üretir."""
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

SCRIPT_DIR = Path(__file__).parent
data = json.load(open(SCRIPT_DIR / 'drc_man_5000_exam_result.json', encoding='utf-8'))
print(f'Yüklenen sonuç: {data["total"]} sınav sorusu')

wb = Workbook()

# Stiller
HF = PatternFill('solid', start_color='1F4E79')
HT = Font(bold=True, color='FFFFFF', size=11, name='Arial')
TF = Font(bold=True, color='1F4E79', size=14, name='Arial')
NF = Font(size=10, name='Arial')
BR = Border(left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'))
GREEN = PatternFill('solid', start_color='D4EDDA')
YELLOW = PatternFill('solid', start_color='FFF3CD')
RED = PatternFill('solid', start_color='F8D7DA')
GRAY = PatternFill('solid', start_color='E9ECEF')

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HT; cell.fill = HF
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BR
    ws.row_dimensions[row].height = 30

def auto_width(ws, max_width=80):
    for col in ws.columns:
        try:
            length = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(length+2, max_width)
        except: pass

# ─── Sayfa 1: Özet ───
ws = wb.active
ws.title = 'Özet'
ws['A1'] = f'DRC MAN 5000-Soru Canlı Sınav Raporu'
ws['A1'].font = TF
ws.merge_cells('A1:D1')
ws['A2'] = f'Tarih: {datetime.now().strftime("%d.%m.%Y %H:%M")} · Süre: {data["elapsed_seconds"]}s · Ortalama: {data["avg_query_ms"]} ms/sorgu'
ws['A2'].font = Font(italic=True, color='666666', name='Arial')

ws['A4'] = 'METRİK'; ws['B4'] = 'DEĞER'; ws['C4'] = 'YÜZDE'; ws['D4'] = 'NOT'
style_header(ws, 4)

rows = [
    ('Toplam Soru', data['total'], '100.0%', ''),
    ('✅ Tam Doğru', data['correct'], f"{data['correct']/data['total']*100:.1f}%", '', GREEN),
    ('🟡 Kısmen Doğru', data['partial'], f"{data['partial']/data['total']*100:.1f}%", '', YELLOW),
    ('❌ Yanlış', data['wrong'], f"{data['wrong']/data['total']*100:.1f}%", '', RED),
    ('⛔ Eşleşmeyen', data['no_match'], f"{data['no_match']/data['total']*100:.1f}%", ''),
    ('🚨 Cost (alış) sızıntısı', data['cost_leaks'], '', '✅ Müşteri korumalı' if data['cost_leaks']==0 else '⚠ DİKKAT'),
    ('📈 Genel Puan', f"{data['score_pct']}%", '', f"Not: {data['grade']}"),
    ('⚡ Performans', f"{data['avg_query_ms']} ms/sorgu", '', f"Toplam {data['elapsed_seconds']}s"),
]
for r, row in enumerate(rows, start=5):
    for c, v in enumerate(row[:4], start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = NF; cell.border = BR
    if len(row) > 4: # color
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = row[4]

# Tip bazlı tablo
ws['A14'] = 'Soru Tipine Göre Performans'
ws['A14'].font = Font(bold=True, size=12, color='1F4E79', name='Arial')
ws['A15'] = 'TİP'; ws['B15'] = 'TOPLAM'; ws['C15'] = 'DOĞRU'; ws['D15'] = 'KISMEN'; ws['E15'] = 'YANLIŞ'; ws['F15'] = 'BAŞARI %'
style_header(ws, 15)
TYPE_LABELS = {'product':'Ürün-spesifik','tech':'Teknik/sorun','calc':'Soğuk Oda Hesabı','mixed':'Kategori/Marka'}
r = 16
for typ, d in data['by_type'].items():
    t = d['c']+d['p']+d['w']
    pct = (d['c']+0.5*d['p'])/t*100 if t else 0
    ws.cell(row=r, column=1, value=TYPE_LABELS.get(typ, typ))
    ws.cell(row=r, column=2, value=t)
    ws.cell(row=r, column=3, value=d['c'])
    ws.cell(row=r, column=4, value=d['p'])
    ws.cell(row=r, column=5, value=d['w'])
    ws.cell(row=r, column=6, value=f'{pct:.1f}%')
    for c in range(1, 7):
        ws.cell(row=r, column=c).font = NF
        ws.cell(row=r, column=c).border = BR
    r += 1

auto_width(ws, max_width=40)

# ─── Sayfa 2: Tüm Sonuçlar ───
ws = wb.create_sheet('Tüm Sonuçlar (5000)')
ws.append(['#', 'Tip', 'Soru', 'Durum', 'Skor', 'Eşleşen Q&A', 'DRC MAN Cevabı', 'Hata Tipi', 'Cost Leak'])
style_header(ws, 1)
for r in data['results']:
    cost_leak_str = '⚠' if r.get('cost_leak') else ''
    ws.append([
        r['no'], TYPE_LABELS.get(r['tip'], r['tip']),
        r['soru'], r['durum'], r['skor'],
        r['eşleşen_q'], r['cevap'], r['hata_tipi'] or '', cost_leak_str
    ])

# Status'a göre renklendir
for row_idx in range(2, ws.max_row+1):
    status = ws.cell(row=row_idx, column=4).value
    fill = None
    if status == 'DOĞRU': fill = GREEN
    elif status == 'KISMEN': fill = YELLOW
    elif status in ('YANLIŞ', 'NO_MATCH'): fill = RED
    if fill:
        for c in range(1, 10):
            ws.cell(row=row_idx, column=c).fill = fill
    for c in range(1, 10):
        ws.cell(row=row_idx, column=c).font = NF
        ws.cell(row=row_idx, column=c).border = BR
        ws.cell(row=row_idx, column=c).alignment = Alignment(wrap_text=True, vertical='top')

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 50
ws.column_dimensions['G'].width = 80
ws.column_dimensions['H'].width = 18
ws.column_dimensions['I'].width = 10
ws.freeze_panes = 'A2'

# ─── Sayfa 3: HATALAR (Yanlış cevaplar) ───
ws = wb.create_sheet('Hatalar')
ws['A1'] = 'YANLIŞ ve EŞLEŞMEYEN Cevaplar — Detaylı Analiz'
ws['A1'].font = TF
ws.merge_cells('A1:H1')
ws.append([])
ws.append(['#', 'Tip', 'Soru', 'Hata Tipi', 'Skor', 'Eşleşen Q&A (yanlış)', 'Verilen Cevap', 'Beklenen ID/Marka'])
style_header(ws, 3)

errors = [r for r in data['results'] if r['durum'] in ('YANLIŞ','NO_MATCH')]
print(f'Hata sayısı: {len(errors)}')
for r in errors:
    expected = ''
    if r.get('beklenen_id'):
        expected = f"#{r['beklenen_id']} (marka: {r.get('beklenen_marka','')}) sale: €{r.get('beklenen_sale',0)}"
    ws.append([
        r['no'], TYPE_LABELS.get(r['tip'], r['tip']),
        r['soru'], r['hata_tipi'], r['skor'],
        r['eşleşen_q'], r['cevap'], expected
    ])

for row_idx in range(4, ws.max_row+1):
    for c in range(1, 9):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = NF; cell.border = BR
        cell.fill = RED
        cell.alignment = Alignment(wrap_text=True, vertical='top')

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 45
ws.column_dimensions['G'].width = 70
ws.column_dimensions['H'].width = 35
ws.freeze_panes = 'A4'

# ─── Sayfa 4: Malzeme Hataları (özel filtre — ürün soruları yanlış olanlar) ───
ws = wb.create_sheet('Malzeme Hataları')
ws['A1'] = 'ÜRÜN/MALZEME SORULARI — Yanlış Eşleşmeler (kullanıcı isteği)'
ws['A1'].font = TF
ws.merge_cells('A1:G1')
ws.append([])
ws.append(['#', 'Soru (müşteri)', 'Beklenen Ürün', 'Beklenen Marka', 'Beklenen Sale', 'Yanlış Match', 'Yanlış Cevap'])
style_header(ws, 3)

material_errors = [r for r in data['results'] if r['tip']=='product' and r['durum'] in ('YANLIŞ','NO_MATCH')]
print(f'Malzeme hatası: {len(material_errors)}')
for r in material_errors:
    ws.append([
        r['no'], r['soru'],
        f"#{r.get('beklenen_id','?')}",
        r.get('beklenen_marka','') or '',
        f"€{r.get('beklenen_sale',0)}" if r.get('beklenen_sale') else '',
        r['eşleşen_q'], r['cevap']
    ])

for row_idx in range(4, ws.max_row+1):
    for c in range(1, 8):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = NF; cell.border = BR
        cell.fill = RED
        cell.alignment = Alignment(wrap_text=True, vertical='top')

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 50
ws.column_dimensions['G'].width = 70
ws.freeze_panes = 'A4'

# ─── Sayfa 5: Doğru Yanıtlar (örnek 200 sample) ───
ws = wb.create_sheet('Doğru Yanıt Örnekleri')
ws['A1'] = 'DOĞRU CEVAP ÖRNEKLERİ — 200 örnek (random)'
ws['A1'].font = TF
ws.merge_cells('A1:F1')
ws.append([])
ws.append(['#', 'Tip', 'Soru', 'Skor', 'Eşleşen Q&A', 'DRC MAN Cevabı'])
style_header(ws, 3)

import random
random.seed(2026)
correct_results = [r for r in data['results'] if r['durum']=='DOĞRU']
sample_correct = random.sample(correct_results, min(200, len(correct_results)))
for r in sample_correct:
    ws.append([r['no'], TYPE_LABELS.get(r['tip'], r['tip']), r['soru'], r['skor'], r['eşleşen_q'], r['cevap']])

for row_idx in range(4, ws.max_row+1):
    for c in range(1, 7):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = NF; cell.border = BR
        cell.fill = GREEN
        cell.alignment = Alignment(wrap_text=True, vertical='top')

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 45
ws.column_dimensions['F'].width = 70
ws.freeze_panes = 'A4'

# Kaydet
out = Path('/Users/anilakbas/Desktop/DRC_MAN_5000_Sinav_Raporu.xlsx')
wb.save(out)
size_mb = out.stat().st_size / 1024 / 1024
print(f'\n✅ Excel raporu üretildi: {out}')
print(f'   Boyut: {size_mb:.2f} MB')
print(f'   Sayfalar: Özet · Tüm Sonuçlar (5000) · Hatalar · Malzeme Hataları · Doğru Örnekleri')
