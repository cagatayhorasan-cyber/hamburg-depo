from __future__ import annotations

import csv
import json
import re
import subprocess
import unicodedata
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/Users/anilakbas/Desktop/Hamburg depo stok programı ")
REPORTS = ROOT / "data/reports"

SOURCE_FILES = [
    Path("/Users/anilakbas/Desktop/Burak2 /CİHAZ SEÇİMLER.xlsx"),
    Path("/Users/anilakbas/Desktop/Burak2 /DANFOSS HERMETİK SERİ-2025.xlsx"),
    Path("/Users/anilakbas/Desktop/Burak2 /DANFOSS SCROLL FRİGOCRAFT  SESSİZ SERİ-2025.xlsx"),
    Path("/Users/anilakbas/Desktop/Burak2 /DANFOSS SCROLL FRİGOCRAFT SERİ-2025.xlsx"),
    Path("/Users/anilakbas/Desktop/Burak2 /DANFOSS SCROLL MERKEZİ FİYAT LİSTESİ-2025.xlsx"),
    Path("/Users/anilakbas/Desktop/Burak2 /ENDÜSTRİYEL SPLİT FİYAT LİSTESİ-2025-REV1.xlsx"),
    Path("/Users/anilakbas/Desktop/Burak2 /Pratik Soğuk Oda Kapasite Belirleme.xlsx"),
    Path("/Users/anilakbas/Desktop/Burak2 /STANDART EVAP. FİYAT LİSTESİ-2025.xlsx"),
    Path("/Users/anilakbas/Desktop/Burak2 /YARI HERMETİK ŞOK CİHAZ SERİSİ -2025.xlsx"),
    Path("/Users/anilakbas/Desktop/Burak2 /cantas_products.csv"),
    Path("/Users/anilakbas/Desktop/Burak2 /cantas_products.xlsx"),
    Path("/Users/anilakbas/Desktop/Burak2 /drc_master.db"),
    Path("/Users/anilakbas/Desktop/Burak2 /fiyat_listesi.xlsx"),
    Path("/Users/anilakbas/Desktop/Burak2 /yeni_liste.xlsx"),
    Path("/Users/anilakbas/Desktop/IBS/Kondenser Ünitesi Fiyat Listesi v1.3.pdf"),
    Path("/Users/anilakbas/Desktop/Panel/turkiye-fiyat-listesi.xlsx"),
    Path("/Users/anilakbas/Desktop/Yılmaz Abi/Hamburg_Stok_Listesi_TekSayfa.xlsx"),
    Path("/Users/anilakbas/Downloads/Hamburg_Stok_Listesi_Full_Final_v2.xlsx"),
    Path("/Users/anilakbas/Downloads/malzeme-listesi.csv"),
    Path("/Users/anilakbas/Desktop/Secop/Slve18CN.xlsx"),
]


def slug(value: str) -> str:
    value = str(value or "").lower()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = re.sub(r"[^a-z0-9]+", "_", value).strip("_")
    return value


def clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def parse_float(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value)
    text = text.replace("€", "").replace("EUR", "").replace("Watt", "").replace("watt", "")
    text = text.replace("W", "").replace("HP", "").replace("V", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    m = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(m.group(0)) if m else None


def detect_brand(*values) -> str:
    text = " ".join(clean_text(v) for v in values).lower()
    for brand in [
        "frigocraft",
        "danfoss",
        "bitzer",
        "dorin",
        "frascold",
        "gea bock",
        "bock",
        "tecumseh",
        "secop",
        "embraco",
        "weiguang",
        "sanhua",
        "alkaterhm",
        "alkatherm",
        "baticompos",
    ]:
        if brand in text:
            return brand.title()
    return ""


def new_record(file: Path, **kwargs):
    base = {
        "source_file": str(file),
        "source_name": file.name,
        "source_sheet": "",
        "source_section": "",
        "source_row": "",
        "record_type": "",
        "code": "",
        "code_role": "",
        "parent_code": "",
        "brand": "",
        "product_name": "",
        "model": "",
        "compressor_brand": "",
        "compressor_model": "",
        "category": "",
        "regime": "",
        "hp": None,
        "voltage": "",
        "refrigerant": "",
        "capacity_w": None,
        "capacity_condition": "",
        "price_eur": None,
        "cost_eur": None,
        "quantity": None,
        "unit": "",
        "stock_quantity": None,
        "stock_unit": "",
        "description": "",
        "raw": "",
    }
    base.update(kwargs)
    return base


def iter_nonempty_rows(ws, start=1):
    for idx, row in enumerate(ws.iter_rows(min_row=start, values_only=True), start=start):
        vals = [clean_text(v) for v in row]
        if any(vals):
            yield idx, vals


def parse_multi_code_sheet(file: Path, ws, doc_family: str):
    records = []
    section = ""
    header_row = None
    for row_idx, vals in iter_nonempty_rows(ws, 1):
        if len(vals) >= 2 and "KOD" in vals[0]:
            header_row = row_idx
            continue
        if header_row is None:
            if vals[0] and not any("KOD" in v for v in vals):
                section = vals[0]
            continue
        if not vals[0] and len(vals) < 2:
            continue
        if vals[0] and "SERİ" in vals[0]:
            section = vals[0]
            continue
        if not vals[0]:
            continue
        # Hermetik/scroll list layouts
        if len(vals) >= 16:
            outdoor_code = clean_text(vals[0])
            if (
                not outdoor_code
                or "KOD" in outdoor_code
                or '"' in outdoor_code
                or "EKİPMAN" in outdoor_code.upper()
                or "LISTESI" in slug(outdoor_code)
                or not re.search(r"\d", outdoor_code)
            ):
                continue
            outdoor_name = clean_text(vals[1])
            if doc_family == "endustriyel_split":
                outdoor_capacity = parse_float(vals[5] or vals[4])
                outdoor_price = parse_float(vals[8])
                indoor_code_idx = 9
                indoor_name_idx = 10
                indoor_qty_idx = 11
                indoor_capacity_idx = 12
                indoor_price_idx = 13
                digital_code_idx = 14
                digital_name_idx = 15
                digital_qty_idx = 16
                digital_price_idx = 17
            else:
                outdoor_capacity = parse_float(vals[4])
                outdoor_price = parse_float(vals[5])
                indoor_code_idx = 6
                indoor_name_idx = 7
                indoor_qty_idx = 8
                indoor_capacity_idx = 9
                indoor_price_idx = 10
                digital_code_idx = 11
                digital_name_idx = 12
                digital_qty_idx = 13
                digital_price_idx = 14
            outdoor = new_record(
                file,
                source_sheet=ws.title,
                source_section=section,
                source_row=row_idx,
                record_type="price_item",
                code=outdoor_code,
                code_role="outdoor_unit",
                brand=detect_brand(outdoor_name, doc_family),
                product_name=outdoor_name,
                model=outdoor_name,
                category=doc_family,
                regime=section,
                hp=parse_float(vals[2]),
                voltage=clean_text(vals[3]),
                capacity_w=outdoor_capacity,
                capacity_condition="-10/+45" if "-10" in clean_text(vals[4]) or doc_family != "endustriyel_split" else "",
                price_eur=outdoor_price,
                raw=" | ".join(v for v in vals[:16] if v),
            )
            records.append(outdoor)

            indoor_code = clean_text(vals[indoor_code_idx]) if len(vals) > indoor_code_idx else ""
            if indoor_code:
                records.append(
                    new_record(
                        file,
                        source_sheet=ws.title,
                        source_section=section,
                        source_row=row_idx,
                        record_type="component_item",
                        code=indoor_code,
                        code_role="indoor_unit",
                        parent_code=outdoor_code,
                        brand=detect_brand(vals[indoor_name_idx], doc_family),
                        product_name=clean_text(vals[indoor_name_idx]),
                        model=clean_text(vals[indoor_name_idx]),
                        category="indoor_unit",
                        regime=section,
                        quantity=parse_float(vals[indoor_qty_idx]),
                        capacity_w=parse_float(vals[indoor_capacity_idx]),
                        price_eur=parse_float(vals[indoor_price_idx]),
                        raw=" | ".join(v for v in vals[:16] if v),
                    )
                )

            # some sheets have digital/controller as the last code block
            digital_code = clean_text(vals[digital_code_idx]) if len(vals) > digital_code_idx else ""
            if digital_code:
                records.append(
                    new_record(
                        file,
                        source_sheet=ws.title,
                        source_section=section,
                        source_row=row_idx,
                        record_type="component_item",
                        code=digital_code,
                        code_role="digital_controller",
                        parent_code=outdoor_code,
                        brand=detect_brand(vals[digital_name_idx], "danfoss"),
                        product_name=clean_text(vals[digital_name_idx]),
                        model=clean_text(vals[digital_name_idx]),
                        category="controller",
                        regime=section,
                        quantity=parse_float(vals[digital_qty_idx]),
                        price_eur=parse_float(vals[digital_price_idx]),
                        raw=" | ".join(v for v in vals[:16] if v),
                    )
                )
        elif len(vals) >= 12:
            # yeni_liste style
            code = clean_text(vals[1])
            if not code or "KOD" in code:
                continue
            outdoor_name = clean_text(vals[2])
            records.append(
                new_record(
                    file,
                    source_sheet=ws.title,
                    source_row=row_idx,
                    source_section=section,
                    record_type="price_item",
                    code=code,
                    code_role="outdoor_unit",
                    brand=detect_brand(vals[0], outdoor_name),
                    product_name=outdoor_name,
                    model=clean_text(vals[3]),
                    compressor_brand=clean_text(vals[0]),
                    compressor_model=clean_text(vals[3]),
                    category=doc_family,
                    hp=parse_float(vals[4]),
                    capacity_w=parse_float(vals[5]),
                    capacity_condition="-10/+40",
                    price_eur=parse_float(vals[6]),
                    raw=" | ".join(v for v in vals[:12] if v),
                )
            )
            indoor_code = clean_text(vals[7])
            if indoor_code:
                records.append(
                    new_record(
                        file,
                        source_sheet=ws.title,
                        source_row=row_idx,
                        source_section=section,
                        record_type="component_item",
                        code=indoor_code,
                        code_role="indoor_unit",
                        parent_code=code,
                        brand=detect_brand(vals[8], outdoor_name),
                        product_name=clean_text(vals[8]),
                        model=clean_text(vals[8]),
                        category="indoor_unit",
                        quantity=parse_float(vals[9]),
                        capacity_w=parse_float(vals[10]),
                        price_eur=parse_float(vals[11]),
                        raw=" | ".join(v for v in vals[:12] if v),
                    )
                )
    return records


def parse_simple_price_sheet(file: Path, ws, kind: str):
    records = []
    header = None
    for row_idx, vals in iter_nonempty_rows(ws, 1):
        if header is None and any("Kodu" in v or "KODU" in v or "EVAP KODU" in v for v in vals):
            header = vals
            continue
        if header is None:
            continue
        first = clean_text(vals[0])
        if not first or "KODU" in first or "Ürün Kodu" == first:
            continue
        if kind == "evaporator":
            if len(vals) < 3:
                continue
            if len(vals) == 3 and parse_float(vals[2]) is not None:
                product_name = clean_text(vals[1])
                description = ""
            else:
                product_name = clean_text(vals[2])
                description = clean_text(vals[1])
            capacity_value = None
            for idx in (4, 3, 5, 6):
                if len(vals) > idx and clean_text(vals[idx]):
                    capacity_value = vals[idx]
                    break
            price_value = vals[-1] if vals else None
            records.append(
                new_record(
                    file,
                    source_sheet=ws.title,
                    source_row=row_idx,
                    record_type="price_item",
                    code=first,
                    code_role="evaporator",
                    brand=detect_brand(product_name, "FrigoCraft"),
                    product_name=product_name,
                    model=product_name,
                    category="evaporator",
                    capacity_w=parse_float(capacity_value),
                    price_eur=parse_float(price_value),
                    description=description,
                    raw=" | ".join(v for v in vals if v),
                )
            )
        else:
            records.append(
                new_record(
                    file,
                    source_sheet=ws.title,
                    source_row=row_idx,
                    record_type="price_item",
                    code=first,
                    code_role="product",
                    brand=detect_brand(vals[1]),
                    product_name=clean_text(vals[1]),
                    model=clean_text(vals[1]),
                    category=kind,
                    price_eur=parse_float(vals[2]) if len(vals) > 2 else None,
                    cost_eur=parse_float(vals[2]) if "maliyet" in slug(" ".join(header)) else None,
                    raw=" | ".join(v for v in vals if v),
                )
            )
    return records


def parse_cihaz_secimleri(file: Path, ws):
    records = []
    header = None
    for row_idx, vals in iter_nonempty_rows(ws, 1):
        if header is None and vals[:3] == ["MAHAL", "HACİM (M3)", "KOD"]:
            header = vals
            continue
        if header is None:
            continue
        code = clean_text(vals[2])
        if not code:
            continue
        records.append(
            new_record(
                file,
                source_sheet=ws.title,
                source_row=row_idx,
                record_type="selection_item",
                code=code,
                code_role="selected_unit",
                brand=detect_brand(vals[3]),
                product_name=clean_text(vals[3]),
                model=clean_text(vals[3]),
                category="selection",
                regime=clean_text(vals[4]),
                hp=parse_float(vals[5]),
                capacity_w=parse_float(vals[6]),
                price_eur=parse_float(vals[8]),
                quantity=parse_float(vals[9]),
                description=f"Mahal: {clean_text(vals[0])} | Hacim: {clean_text(vals[1])}",
                raw=" | ".join(v for v in vals if v),
            )
        )
    return records


def parse_cantas_xlsx(file: Path, ws):
    records = []
    header = None
    for row_idx, vals in iter_nonempty_rows(ws, 1):
        if header is None and vals[:2] == ["Ürün Adı", "Fiyat (EUR)"]:
            header = vals
            continue
        if header is None:
            continue
        name = clean_text(vals[0])
        if not name:
            continue
        code = clean_text(vals[3]) or name
        records.append(
            new_record(
                file,
                source_sheet=ws.title,
                source_row=row_idx,
                record_type="technical_price_item",
                code=code,
                code_role="catalog_model",
                brand=clean_text(vals[2]),
                product_name=name,
                model=clean_text(vals[3]),
                category=clean_text(vals[4] or ws.title),
                regime=clean_text(vals[5]),
                hp=parse_float(vals[6]),
                voltage=clean_text(vals[7]),
                refrigerant=clean_text(vals[8]),
                capacity_w=parse_float(vals[9]),
                price_eur=parse_float(vals[1]),
                description=clean_text(vals[10] if len(vals) > 10 else ""),
                raw=" | ".join(v for v in vals if v),
            )
        )
    return records


def parse_cantas_csv(file: Path):
    records = []
    with open(file, "r", encoding="utf-8", errors="ignore") as fh:
        reader = csv.DictReader(fh)
        for idx, row in enumerate(reader, start=2):
            name = clean_text(row.get("name"))
            if not name:
                continue
            records.append(
                new_record(
                    file,
                    source_row=idx,
                    record_type="price_item",
                    code=name,
                    code_role="catalog_name",
                    brand=detect_brand(name),
                    product_name=name,
                    model=name,
                    category="catalog_csv",
                    price_eur=parse_float(row.get("price")),
                    raw=json.dumps(row, ensure_ascii=False),
                )
            )
    return records


def parse_stock_xlsx(file: Path, ws, full=False):
    records = []
    for row_idx, vals in iter_nonempty_rows(ws, 2):
        category = clean_text(vals[0])
        name = clean_text(vals[1])
        qty = clean_text(vals[2])
        if not name:
            continue
        m = re.search(r"(\d+(?:[.,]\d+)?)", qty)
        stock_quantity = parse_float(m.group(1)) if m else parse_float(qty)
        stock_unit = clean_text(vals[3]) if full and len(vals) > 3 else re.sub(r"[\d\s.,]+", "", qty).strip()
        records.append(
            new_record(
                file,
                source_sheet=ws.title,
                source_row=row_idx,
                record_type="stock_observation",
                code=name,
                code_role="stock_name",
                brand=detect_brand(name),
                product_name=name,
                model=name,
                category=category,
                stock_quantity=stock_quantity,
                stock_unit=stock_unit,
                raw=" | ".join(v for v in vals if v),
            )
        )
    return records


def parse_malzeme_csv(file: Path):
    records = []
    with open(file, "r", encoding="utf-8-sig", errors="ignore") as fh:
        reader = csv.reader(fh)
        rows = list(reader)
    for idx, row in enumerate(rows[2:], start=3):
        if len(row) < 4:
            continue
        category, name, qty, unit = (clean_text(x) for x in row[:4])
        if not name:
            continue
        records.append(
            new_record(
                file,
                source_row=idx,
                record_type="stock_observation",
                code=name,
                code_role="stock_name",
                brand=detect_brand(name),
                product_name=name,
                model=name,
                category=category,
                stock_quantity=parse_float(qty),
                stock_unit=unit,
                description=clean_text(row[4]) if len(row) > 4 else "",
                raw=" | ".join(clean_text(v) for v in row if clean_text(v)),
            )
        )
    return records


def parse_panel_xlsx(file: Path, ws):
    records = []
    for row_idx, vals in iter_nonempty_rows(ws, 5):
        name = clean_text(vals[1] if len(vals) > 1 else "")
        if not name or name == "Ürün":
            continue
        records.append(
            new_record(
                file,
                source_sheet=ws.title,
                source_row=row_idx,
                record_type="price_item",
                code=name,
                code_role="panel_product",
                brand="Baticompos France",
                product_name=name,
                model=name,
                category="panel",
                unit=clean_text(vals[3]) if len(vals) > 3 else "",
                price_eur=parse_float(vals[4]) if len(vals) > 4 else None,
                description=clean_text(vals[5]) if len(vals) > 5 else "",
                raw=" | ".join(v for v in vals if v),
            )
        )
    return records


def parse_ibs_pdf(file: Path):
    text = subprocess.run(
        ["pdftotext", "-f", "1", "-l", "6", "-nopgbrk", str(file), "-"],
        stdout=subprocess.PIPE,
        stderr=subprocess.DEVNULL,
        text=True,
        timeout=10,
    ).stdout
    lines = [clean_text(line) for line in text.splitlines() if clean_text(line)]
    records = []
    for idx, line in enumerate(lines):
        if re.match(r"^ICF-[A-Z0-9.-]+$", line):
            code = line
            price = None
            compressor_model = ""
            product_name = ""
            for back in range(1, 8):
                prev = lines[idx - back]
                if "€" in prev:
                    price = parse_float(prev)
                    continue
                if re.search(r"[A-Z]{2,}\s?\d", prev) and not compressor_model:
                    compressor_model = prev
            product_name = f"Tecumseh Microchannel Kondenser Ünitesi {compressor_model}".strip()
            records.append(
                new_record(
                    file,
                    source_row=idx + 1,
                    record_type="price_item",
                    code=code,
                    code_role="condensing_unit",
                    brand="Tecumseh",
                    product_name=product_name,
                    model=code,
                    compressor_brand="Tecumseh",
                    compressor_model=compressor_model,
                    category="microchannel_condensing_unit",
                    price_eur=price,
                    description="IBS Tecumseh kondenser ünitesi fiyat listesi",
                    raw=code,
                )
            )
    return records


def parse_secop_xlsx(file: Path):
    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    records = []
    with zipfile.ZipFile(file) as zf:
        sst = []
        root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
        for si in root.findall("a:si", ns):
            sst.append("".join(t.text or "" for t in si.iterfind(".//a:t", ns)))
        sheet = ET.fromstring(zf.read("xl/worksheets/sheet1.xml"))
        rows = []
        for row in sheet.findall(".//a:row", ns):
            values = {}
            for cell in row.findall("a:c", ns):
                ref = cell.attrib.get("r")
                typ = cell.attrib.get("t")
                val = cell.find("a:v", ns)
                if val is None:
                    continue
                text = sst[int(val.text)] if typ == "s" else val.text
                values[ref] = clean_text(text)
            if values:
                rows.append(values)
    for row in rows:
        if row.get("F23") and row.get("G23"):
            break
    for values in rows:
        if values.get("F23"):  # not used, keep lint quiet
            pass
    for row in rows:
        code = row.get("F23") if False else None
    for row in rows:
        code = row.get("F23")
        if code:
            pass
    for row in rows:
        sales = row.get("F23") or row.get("F24") or row.get("F25") or row.get("F26")
    # data rows are B23:AG26 style
    for row in rows:
        sales = row.get("F23") or row.get("F24") or row.get("F25") or row.get("F26")
    for row in rows:
        sales = next((v for k, v in row.items() if k.startswith("F")), "")
        desc = next((v for k, v in row.items() if k.startswith("G")), "")
        refr = next((v for k, v in row.items() if k.startswith("H")), "")
        if not sales or sales == "Sales number":
            continue
        records.append(
            new_record(
                file,
                source_sheet="Sayfa1",
                source_row=next((k for k, v in row.items() if k.startswith("F")), ""),
                record_type="technical_point",
                code=sales,
                code_role="sales_number",
                brand="Secop",
                product_name=desc,
                model=desc,
                category="compressor_datasheet",
                refrigerant=refr,
                voltage=next((v for k, v in row.items() if k.startswith("I")), ""),
                capacity_w=parse_float(next((v for k, v in row.items() if k.startswith("S")), None)),
                capacity_condition=f"pevap {next((v for k, v in row.items() if k.startswith('B')), '')} | pcond {next((v for k, v in row.items() if k.startswith('C')), '')} | Tamb {next((v for k, v in row.items() if k.startswith('AG')), '')}",
                description=f"Hz {next((v for k, v in row.items() if k.startswith('J')), '')} | Application {next((v for k, v in row.items() if k.startswith('P')), '')} | Status {next((v for k, v in row.items() if k.startswith('R')), '')}",
                raw=json.dumps(row, ensure_ascii=False),
            )
        )
    return records


def parse_drc_master_db(file: Path):
    import sqlite3

    records = []
    conn = sqlite3.connect(file)
    cur = conn.cursor()
    for idx, row in enumerate(cur.execute("select id,name,category,unit_price,cooling_capacity,regime from products"), start=1):
        records.append(
            new_record(
                file,
                source_row=idx,
                record_type="db_product",
                code=f"DB-PRODUCT-{row[0]}",
                code_role="db_id",
                brand=detect_brand(row[1]),
                product_name=clean_text(row[1]),
                model=clean_text(row[1]),
                category=clean_text(row[2]),
                regime=clean_text(row[5]),
                price_eur=parse_float(row[3]),
                capacity_w=parse_float(row[4]),
                raw=json.dumps(row, ensure_ascii=False),
            )
        )
    conn.close()
    return records


def parse_file(file: Path):
    docs = []
    records = []
    suffix = file.suffix.lower()
    if suffix == ".xlsx":
        wb = load_workbook(file, read_only=True, data_only=True)
        docs.append(
            {
                "file": str(file),
                "name": file.name,
                "type": "xlsx",
                "sheet_names": wb.sheetnames,
                "size_kb": round(file.stat().st_size / 1024, 1),
            }
        )
        for ws in wb.worksheets:
            title = ws.title
            if file.name.startswith("CI"):
                records.extend(parse_cihaz_secimleri(file, ws))
            elif "HERMETI" in file.name or "SCROLL" in file.name:
                if title in {"Sayfa1", "Sayfa2", "LIST"}:
                    records.extend(parse_simple_price_sheet(file, ws, "price_list"))
                else:
                    records.extend(parse_multi_code_sheet(file, ws, slug(file.stem)))
            elif "ENDU" in file.name or file.name == "fiyat_listesi.xlsx":
                if title == "Sayfa1":
                    records.extend(parse_simple_price_sheet(file, ws, "endustriyel_split"))
                else:
                    records.extend(parse_multi_code_sheet(file, ws, "endustriyel_split"))
            elif "YARI HERMET" in file.name:
                if title.startswith("Sayfa"):
                    records.extend(parse_simple_price_sheet(file, ws, "shock_unit"))
                else:
                    records.extend(parse_multi_code_sheet(file, ws, "shock_unit"))
            elif "STANDART EVAP" in file.name:
                records.extend(parse_simple_price_sheet(file, ws, "evaporator"))
            elif file.name == "cantas_products.xlsx":
                if title == "Kompresörler":
                    records.extend(parse_cantas_xlsx(file, ws))
            elif "Pratik" in file.name:
                # calculator reference, keep doc only
                pass
            elif file.name == "yeni_liste.xlsx":
                records.extend(parse_multi_code_sheet(file, ws, "endustriyel_split"))
            elif file.name == "turkiye-fiyat-listesi.xlsx":
                records.extend(parse_panel_xlsx(file, ws))
            elif "Hamburg_Stok_Listesi_TekSayfa" in file.name:
                records.extend(parse_stock_xlsx(file, ws, full=False))
            elif "Hamburg_Stok_Listesi_Full_Final_v2" in file.name:
                records.extend(parse_stock_xlsx(file, ws, full=True))
            elif file.name == "Slve18CN.xlsx":
                records.extend(parse_secop_xlsx(file))
    elif suffix == ".csv":
        docs.append(
            {
                "file": str(file),
                "name": file.name,
                "type": "csv",
                "size_kb": round(file.stat().st_size / 1024, 1),
            }
        )
        if file.name == "cantas_products.csv":
            records.extend(parse_cantas_csv(file))
        elif file.name == "malzeme-listesi.csv":
            records.extend(parse_malzeme_csv(file))
    elif suffix == ".pdf":
        docs.append(
            {
                "file": str(file),
                "name": file.name,
                "type": "pdf",
                "size_kb": round(file.stat().st_size / 1024, 1),
            }
        )
        if "Kondenser" in file.name:
            records.extend(parse_ibs_pdf(file))
    elif suffix == ".db":
        docs.append(
            {
                "file": str(file),
                "name": file.name,
                "type": "sqlite",
                "size_kb": round(file.stat().st_size / 1024, 1),
            }
        )
        records.extend(parse_drc_master_db(file))
    return docs, records


all_docs = []
all_records = []
for source in SOURCE_FILES:
    docs, records = parse_file(source)
    all_docs.extend(docs)
    all_records.extend(records)

# normalize blank codes out
all_records = [r for r in all_records if clean_text(r["code"])]

master = defaultdict(lambda: {
    "code": "",
    "roles": set(),
    "brands": set(),
    "names": set(),
    "models": set(),
    "categories": set(),
    "source_files": set(),
    "prices": set(),
    "costs": set(),
    "capacities": set(),
    "refrigerants": set(),
    "parent_codes": set(),
    "stock_quantities": [],
    "record_count": 0,
})

for record in all_records:
    m = master[record["code"]]
    m["code"] = record["code"]
    m["roles"].add(record["code_role"])
    if record["brand"]:
        m["brands"].add(record["brand"])
    if record["product_name"]:
        m["names"].add(record["product_name"])
    if record["model"]:
        m["models"].add(record["model"])
    if record["category"]:
        m["categories"].add(record["category"])
    m["source_files"].add(record["source_name"])
    if record["price_eur"] is not None:
        m["prices"].add(round(record["price_eur"], 4))
    if record["cost_eur"] is not None:
        m["costs"].add(round(record["cost_eur"], 4))
    if record["capacity_w"] is not None:
        m["capacities"].add(round(record["capacity_w"], 4))
    if record["refrigerant"]:
        m["refrigerants"].add(record["refrigerant"])
    if record["parent_code"]:
        m["parent_codes"].add(record["parent_code"])
    if record["stock_quantity"] is not None:
        m["stock_quantities"].append({"qty": record["stock_quantity"], "unit": record["stock_unit"], "source": record["source_name"]})
    m["record_count"] += 1

master_rows = []
conflicts = []
for code, data in sorted(master.items()):
    row = {
        "code": code,
        "roles": sorted(data["roles"]),
        "brands": sorted(data["brands"]),
        "names": sorted(data["names"])[:8],
        "models": sorted(data["models"])[:8],
        "categories": sorted(data["categories"]),
        "source_files": sorted(data["source_files"]),
        "prices": sorted(data["prices"]),
        "costs": sorted(data["costs"]),
        "capacities": sorted(data["capacities"]),
        "refrigerants": sorted(data["refrigerants"]),
        "parent_codes": sorted(data["parent_codes"]),
        "stock_quantities": data["stock_quantities"],
        "record_count": data["record_count"],
    }
    master_rows.append(row)
    if len(row["names"]) > 1 or len(row["prices"]) > 1 or len(row["capacities"]) > 1:
        conflicts.append(
            {
                "code": code,
                "name_count": len(row["names"]),
                "price_count": len(row["prices"]),
                "capacity_count": len(row["capacities"]),
                "names": row["names"],
                "prices": row["prices"],
                "capacities": row["capacities"],
                "source_files": row["source_files"],
            }
        )

REPORTS.mkdir(parents=True, exist_ok=True)
json_path = REPORTS / "kaynak-dokuman-konsolide-2026-04-17.json"
csv_path = REPORTS / "kaynak-dokuman-konsolide-kayitlar-2026-04-17.csv"
master_path = REPORTS / "kaynak-dokuman-kod-master-2026-04-17.json"
conflict_path = REPORTS / "kaynak-dokuman-cakisma-2026-04-17.json"
summary_path = REPORTS / "kaynak-dokuman-ozet-2026-04-17.md"

json_path.write_text(
    json.dumps(
        {
            "generatedAt": __import__("datetime").datetime.now().isoformat(),
            "documentCount": len(all_docs),
            "recordCount": len(all_records),
            "documents": all_docs,
            "records": all_records,
        },
        ensure_ascii=False,
        indent=2,
    )
)

csv_columns = [
    "source_name", "source_sheet", "source_section", "source_row", "record_type", "code", "code_role",
    "parent_code", "brand", "product_name", "model", "compressor_brand", "compressor_model", "category",
    "regime", "hp", "voltage", "refrigerant", "capacity_w", "capacity_condition", "price_eur", "cost_eur",
    "quantity", "unit", "stock_quantity", "stock_unit", "description", "raw",
]
with csv_path.open("w", newline="", encoding="utf-8") as fh:
    writer = csv.DictWriter(fh, fieldnames=csv_columns)
    writer.writeheader()
    for row in all_records:
        writer.writerow({key: row.get(key) for key in csv_columns})

master_path.write_text(json.dumps({"generatedAt": __import__("datetime").datetime.now().isoformat(), "totalCodes": len(master_rows), "rows": master_rows}, ensure_ascii=False, indent=2))
conflict_path.write_text(json.dumps({"generatedAt": __import__("datetime").datetime.now().isoformat(), "totalConflicts": len(conflicts), "rows": conflicts}, ensure_ascii=False, indent=2))

summary = [
    "# Kaynak Dokuman Konsolidasyon Ozeti",
    "",
    f"- Kaynak dokuman: {len(all_docs)}",
    f"- Normalize kayit: {len(all_records)}",
    f"- Tekil kod/master kayit: {len(master_rows)}",
    f"- Cakismanin oldugu kod: {len(conflicts)}",
    "",
    "## Kaynak Dokumanlar",
    "",
]
for doc in all_docs:
    summary.append(f"- {doc['name']} | {doc['type']} | {doc.get('size_kb','')} KB")
summary.extend([
    "",
    "## En Cok Tekrarlanan Kodlar",
    "",
])
for row in sorted(master_rows, key=lambda x: x["record_count"], reverse=True)[:20]:
    summary.append(f"- {row['code']} | kayit: {row['record_count']} | roller: {', '.join(row['roles'])} | dosya: {', '.join(row['source_files'][:4])}")
summary.extend([
    "",
    "## Cakisma Ornekleri",
    "",
])
for row in conflicts[:30]:
    summary.append(f"- {row['code']} | isim:{row['name_count']} fiyat:{row['price_count']} kapasite:{row['capacity_count']} | dosya: {', '.join(row['source_files'][:4])}")
summary_path.write_text("\n".join(summary), encoding="utf-8")

print(
    json.dumps(
        {
            "documentCount": len(all_docs),
            "recordCount": len(all_records),
            "masterCodeCount": len(master_rows),
            "conflictCount": len(conflicts),
            "jsonPath": str(json_path),
            "csvPath": str(csv_path),
            "masterPath": str(master_path),
            "conflictPath": str(conflict_path),
            "summaryPath": str(summary_path),
        },
        ensure_ascii=False,
        indent=2,
    )
)
